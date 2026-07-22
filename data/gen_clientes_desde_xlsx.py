# -*- coding: utf-8 -*-
"""Carga/actualiza la base de clientes (data/clientes.csv) desde la Plantilla-Clientes.xlsx.

Lee las columnas POR NOMBRE de encabezado (Codigo, Nombre) sin importar el orden ni
que haya un titulo arriba. UPSERT: agrega los codigos nuevos y actualiza el nombre de
los que ya existen; NO borra a nadie. Conserva el orden actual y agrega los nuevos al
final. El codigo se rellena a 6 digitos por si Excel le quito los ceros.

Uso:
  python gen_clientes_desde_xlsx.py "<ruta de la Plantilla-Clientes.xlsx>"
Por defecto lee Plantilla-Clientes.xlsx del proyecto y escribe data/clientes.csv.
"""
import openpyxl, csv, unicodedata, re, os, sys

D = os.path.dirname(os.path.abspath(__file__))
APP = os.path.dirname(D)
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(APP, 'Plantilla-Clientes.xlsx')
OUT = sys.argv[2] if len(sys.argv) > 2 else os.path.join(D, 'clientes.csv')


def key(s):
    s = unicodedata.normalize('NFD', str(s or ''))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower()
    return re.sub(r'[^a-z0-9]+', '', s)


ALIAS = {'codigo': ['codigo', 'cod', 'codcliente', 'codigocliente'],
         'nombre': ['nombre', 'nombrecliente', 'cliente', 'razonsocial']}


def norm_cod(v):
    if v is None:
        return ''
    if isinstance(v, float):
        v = int(v)
    s = str(v).strip()
    if not s:
        return ''
    return s.zfill(6) if s.isdigit() and len(s) < 6 else s


def cell_text(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# --- base actual (orden y contenido) ---
orden, base = [], {}
if os.path.exists(OUT):
    with open(OUT, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            cod = norm_cod(row.get('Codigo'))
            if not cod:
                continue
            if cod not in base:
                orden.append(cod)
            base[cod] = (row.get('Nombre') or '').strip()

# --- plantilla ---
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Clientes'] if 'Clientes' in wb.sheetnames else wb.active
filas = list(ws.iter_rows(values_only=True))
if not filas:
    print('La plantilla esta vacia.')
    sys.exit(1)

colmap, hdr = {}, 0
for ri, fila in enumerate(filas[:10]):
    m = {}
    for i, c in enumerate(fila):
        k = key(c)
        for lg, al in ALIAS.items():
            if k in al and lg not in m:
                m[lg] = i
    if len(m) > len(colmap):
        colmap, hdr = m, ri

if 'codigo' not in colmap or 'nombre' not in colmap:
    print('La plantilla debe tener las columnas Codigo y Nombre.')
    sys.exit(1)

nuevos, actualizados, sin_cambio = 0, 0, 0
for r in filas[hdr + 1:]:
    ci, ni = colmap['codigo'], colmap['nombre']
    cod = norm_cod(r[ci] if ci < len(r) else None)
    nom = cell_text(r[ni] if ni < len(r) else None)
    if not cod or not nom:
        continue
    if cod not in base:
        orden.append(cod)
        base[cod] = nom
        nuevos += 1
    elif base[cod] != nom:
        base[cod] = nom
        actualizados += 1
    else:
        sin_cambio += 1

with open(OUT, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['Codigo', 'Nombre'])
    for cod in orden:
        w.writerow([cod, base[cod]])

print('clientes en la base:', len(orden))
print('  nuevos:', nuevos, '| actualizados:', actualizados, '| sin cambio:', sin_cambio)
print('CSV:', OUT)
print('\nSiguiente paso para publicar: python build.py --solo-datos --push')
