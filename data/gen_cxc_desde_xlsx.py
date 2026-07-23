# -*- coding: utf-8 -*-
"""Convierte el xlsx de Cuentas por Cobrar a data/cxc.csv.

Lee las columnas POR NOMBRE de encabezado (no por posicion), asi la plantilla
puede tener las columnas en cualquier orden. Usa la Plantilla-CxC.xlsx del
proyecto (ver ReporteCobranza/Plantilla-CxC.xlsx).

Campos que llena el usuario (encabezados aceptados, sin importar mayus/acentos):
  FechaFactura   (o "Fecha de la Factura", "Fecha Factura")      dd/mm/aaaa
  Factura        (o "Numero de Factura", "N Factura", "Nro")
  Tipo           (o "Tipo de Documento")
  Cliente        (o "Nombre", "Nombre Cliente")
  Codigo         (OPCIONAL: si se deja vacio, se resuelve por el nombre)
  Vencimiento    (o "Fecha de Vencimiento")                       dd/mm/aaaa
  Base           (o "Base $")                                     monto en $
  IVA            (o "IVA $")                                       monto en $
  TotalPorPagar  (o "Total por Pagar", "Saldo")                   monto en $
  Observacion    (o "Observaciones")                              texto (opcional)

Campos que se calculan solos (NO se llenan en la plantilla):
  TasaFact  = tasa BCV del dia de la fecha de la factura (del historico)
  IVABs     = IVA $ x TasaFact
  IVA25     = IVABs x 0.25

Uso:
  python gen_cxc_desde_xlsx.py "<ruta del cxc.xlsx>" [salida.csv]
Por defecto lee Plantilla-CxC.xlsx del proyecto y escribe data/cxc.csv.
"""
import openpyxl, csv, unicodedata, re, os, sys, datetime, collections

D = os.path.dirname(os.path.abspath(__file__))
APP = os.path.dirname(D)
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(APP, 'Plantilla-CxC.xlsx')
OUT = sys.argv[2] if len(sys.argv) > 2 else os.path.join(D, 'cxc.csv')
CLI = os.path.join(D, 'clientes.csv')
TAS = os.path.join(D, 'tasas_bcv.csv')


def norm(s):
    s = unicodedata.normalize('NFD', str(s))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    s = re.sub(r'\b(c a|ca|compania anonima|s a|sa|srl|rl|cia)\b', '', s).strip()
    return re.sub(r'\s+', ' ', s)


def key(s):
    """Encabezado normalizado para emparejar columnas: minusculas, sin acentos ni signos."""
    s = unicodedata.normalize('NFD', str(s or ''))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower()
    return re.sub(r'[^a-z0-9]+', '', s)


# nombre logico de la columna -> lista de encabezados aceptados (ya normalizados con key)
ALIAS = {
    'fecha':   ['fechafactura', 'fechadelafactura', 'fechafact', 'fecha'],
    'factura': ['factura', 'numerodefactura', 'nfactura', 'nrofactura', 'nrodefactura', 'numfactura'],
    'tipo':    ['tipo', 'tipodedocumento', 'tipodocumento'],
    'cliente': ['cliente', 'nombrecliente', 'nombre'],
    'codigo':  ['codigo', 'codcliente', 'codigocliente', 'cod'],
    'venc':    ['vencimiento', 'fechadevencimiento', 'fechavencimiento'],
    'base':    ['base', 'base$', 'basedolares'],
    'iva':     ['iva', 'iva$', 'ivadolares'],
    'total':   ['totalporpagar', 'totalapagar', 'saldo', 'saldopendiente', 'totalpagar'],
    'obs':     ['observacion', 'observaciones', 'obs'],
}


def num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if v is None:
        return None
    s = str(v).strip().replace('.', '').replace(',', '.') if re.search(r',\d{1,2}$', str(v)) else str(v).strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def as_date(v):
    """Devuelve un date desde un datetime o un texto dd/mm/aaaa (o aaaa-mm-dd)."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v or '').strip()
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def d2s(v):
    d = as_date(v)
    return d.strftime('%d/%m/%Y') if d else (str(v).strip() if v else '')


# --- clientes: nombre normalizado -> codigo, y set de codigos validos ---
idx, cods = {}, set()
with open(CLI, encoding='utf-8') as f:
    for row in csv.DictReader(f):
        idx.setdefault(norm(row['Nombre']), row['Codigo'])
        cods.add(row['Codigo'].strip())

# --- tasas: fecha ISO -> USD ---
tasas = {}
with open(TAS, encoding='utf-8') as f:
    for row in csv.DictReader(f):
        if row['USD']:
            tasas[row['Fecha']] = float(row['USD'])

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Facturas'] if 'Facturas' in wb.sheetnames else wb.active

rows_iter = list(ws.iter_rows(values_only=True))
if not rows_iter:
    print('La plantilla esta vacia.')
    sys.exit(1)


# Columnas FIJAS por posicion (A=0). Solo se leen estas; el resto de la plantilla se ignora.
# A=Codigo C=FechaFactura D=Factura E=Tipo N=Cliente T=Vencimiento V=Base W=IVA Z=TotalPorPagar AI=Observacion
POS = {'codigo': 0, 'fecha': 2, 'factura': 3, 'tipo': 4, 'cliente': 13,
       'venc': 19, 'base': 21, 'iva': 22, 'total': 25, 'obs': 34}

# ubicar la fila de encabezado (dice Factura en D o Codigo en A) para saber donde empiezan los datos
hdr_row = -1
for idx_row, fila in enumerate(rows_iter[:15]):
    kA = key(fila[POS['codigo']]) if len(fila) > POS['codigo'] else ''
    kC = key(fila[POS['fecha']]) if len(fila) > POS['fecha'] else ''
    kD = key(fila[POS['factura']]) if len(fila) > POS['factura'] else ''
    if kD == 'factura' or kA == 'codigo' or kC in ('fechafactura', 'fecha'):
        hdr_row = idx_row
        break
# si no se hallo encabezado (hdr_row=-1), rows_iter[hdr_row+1:] = rows_iter[0:] procesa todo


def get(r, logic):
    i = POS.get(logic)
    return r[i] if (i is not None and i < len(r)) else None


rows, sin_cod, sin_tasa = [], collections.Counter(), 0
for r in rows_iter[hdr_row + 1:]:
    if get(r, 'factura') in (None, '') and get(r, 'cliente') in (None, ''):
        continue                                   # fila vacia
    nombre = str(get(r, 'cliente') or '').strip()
    cod = str(get(r, 'codigo') or '').strip()
    if cod and cod not in cods:
        cod = ''                                   # codigo escrito que no existe: se ignora
    if not cod:
        cod = idx.get(norm(nombre), '')
    if not cod:
        sin_cod[nombre] += 1

    fecha = as_date(get(r, 'fecha'))
    iso = fecha.isoformat() if fecha else ''
    tasa = tasas.get(iso)
    if fecha and tasa is None:
        sin_tasa += 1

    base = num(get(r, 'base'))
    iva = num(get(r, 'iva'))
    total = num(get(r, 'total'))
    iva_bs = round(iva * tasa, 2) if (iva is not None and tasa) else None
    iva25 = round(iva_bs * 0.25, 2) if iva_bs is not None else None

    rows.append([
        cod, nombre,
        str(get(r, 'factura') or '').strip(),
        str(get(r, 'tipo') or '').strip() or 'Factura',
        d2s(get(r, 'fecha')),
        round(tasa, 4) if tasa else '',
        d2s(get(r, 'venc')),
        base if base is not None else '',
        iva if iva is not None else '',
        iva_bs if iva_bs is not None else '',
        iva25 if iva25 is not None else '',
        total if total is not None else '',
        str(get(r, 'obs') or '').strip(),
    ])

with open(OUT, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['Codigo', 'Cliente', 'Factura', 'Tipo', 'FechaFactura', 'TasaFact',
                'Vencimiento', 'Base', 'IVA', 'IVABs', 'IVA25', 'TotalPorPagar', 'Observacion'])
    w.writerows(rows)

con = sum(1 for r in rows if r[0])
print('facturas:', len(rows), '| con codigo:', con, '| sin codigo:', len(rows) - con)
print('sin tasa (fecha fuera del historico):', sin_tasa)
if sin_cod:
    print('clientes sin codigo:', len(sin_cod))
    for n, c in sin_cod.most_common(20):
        print('   -', n, '(%d)' % c)
print('CSV:', OUT)
print('\nSiguiente paso para publicar: python build.py --solo-datos --push')
